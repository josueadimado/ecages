import json
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from datetime import timedelta
from django.db.models import Q, Count, Sum, F
from django.db import transaction
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.urls import reverse

from .models import RestockRequest, RestockLine, WarehousePurchaseRequest, WarehousePurchaseLine
from apps.inventory.models import SalesPointStock, StockTransaction, SalesPoint
from apps.products.models import Product
from .models import TransferRequest, TransferRequestLine
@login_required
def warehouse_inbound_cd(request):
    """Approvisionnement (CD -> Entrepôt): demandes en attente à confirmer."""
    role = getattr(request.user, 'role', '')
    if not (request.user.is_superuser or role == 'warehouse_mgr' or getattr(request.user, 'is_staff', False)):
        return redirect('sales:dashboard')

    warehouse = SalesPoint.objects.filter(is_warehouse=True).first()
    qs = (
        RestockRequest.objects.select_related('provider', 'requested_by')
        .filter(salespoint=warehouse, status__in=['sent','partially_validated'])
        .order_by('-created_at')
    )

    q = (request.GET.get('q') or '').strip()
    if q:
        qs = qs.filter(Q(reference__icontains=q))

    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    paginator = Paginator(qs, 25)
    page_number = request.GET.get('page') or 1
    try:
        page_obj = paginator.get_page(page_number)
    except (PageNotAnInteger, EmptyPage):
        page_obj = paginator.get_page(1)

    return render(request, 'inventory/warehouse/warehouse_inbound_cd.html', {
        'rows': page_obj.object_list,
        'page_obj': page_obj,
        'paginator': paginator,
    })

# ===== Warehouse low-stock purchase builder (to Commercial Director) =====

@login_required
def warehouse_purchase_builder(request):
    """Show finishing products in warehouse and allow sending a purchase request to Commercial Director.
    Uses existing API api_wh_cmd_submit to create WarehousePurchaseRequest.
    """
    if not (request.user.is_superuser or getattr(request.user, 'role', '') == 'warehouse_mgr' or getattr(request.user, 'is_staff', False)):
        return redirect('sales:dashboard')

    kind = (request.GET.get('type') or 'piece').strip()
    if kind not in {'piece','moto','all'}:
        kind = 'piece'
    q = (request.GET.get('q') or '').strip()

    # Find warehouse SP
    wh = SalesPoint.objects.filter(is_warehouse=True).first()
    if not wh:
        wh = SalesPoint.objects.filter(Q(name__icontains='entrep') | Q(name__icontains='ware')).order_by('name').first()

    # Build list of low/finished products
    items = []
    base_products = Product.objects.filter(is_active=True)
    if kind != 'all':
        base_products = base_products.filter(product_type=kind)
    products = base_products
    if q:
        products = products.filter(Q(name__icontains=q) | Q(brand__name__icontains=q))

    # Exclude products already in active purchase requests (draft/sent/acknowledged)
    active_product_ids = set(
        WarehousePurchaseLine.objects.filter(
            request__status__in=['draft','sent','acknowledged']
        ).values_list('product_id', flat=True)
    )

    for p in products.select_related('brand'):
        if p.id in active_product_ids:
            continue
        sps = None
        if wh:
            try:
                sps = SalesPointStock.objects.get(salespoint=wh, product=p)
            except SalesPointStock.DoesNotExist:
                sps = None
        remaining = int(getattr(sps, 'remaining_qty', 0) or 0)
        alert = int(getattr(sps, 'alert_qty', 5) or 5)
        # Define low strictly by the product alert threshold
        threshold = alert
        if remaining <= threshold:
            items.append({
                'product': p,
                'remaining_qty': remaining,
                'alert_qty': alert,
                # naive suggested quantity to reach 3x alert
                'suggested_qty': max(0, alert * 3 - remaining) or 1,
            })
            # Count low stock items for possible summary notification later
            try:
                __low_stock_count = (__low_stock_count + 1) if '___low_stock_marker' in globals() else 1
            except Exception:
                __low_stock_count = 1
            globals()['___low_stock_marker'] = True

    # After collecting items, send ONE summary notification for warehouse managers (no per-product spam)
    try:
        if globals().get('___low_stock_marker') and (__low_stock_count or 0) > 0:
            from django.utils import timezone as _tz
            from django.contrib.auth import get_user_model
            from apps.sales.models import Notification
            User = get_user_model()
            today = _tz.localdate()
            msg = f"⚠️ Stock bas (Entrepôt) · {__low_stock_count} article(s) en alerte"
            for u in User.objects.filter(role='warehouse_mgr', is_active=True):
                exists = Notification.objects.filter(user=u, created_at__date=today, kind='low_stock_wh_summary').exists()
                if not exists:
                    Notification.objects.create(user=u, message=msg, link="/inventory/warehouse/purchase/", kind='low_stock_wh_summary')
    except Exception:
        pass
    finally:
        if '___low_stock_marker' in globals():
            try:
                del globals()['___low_stock_marker']
                del __low_stock_count
            except Exception:
                pass

    # Compute stock summary for this kind at warehouse
    total_in_stock = 0
    total_products_all = 0
    low_count = 0
    if wh:
        qs_stats = SalesPointStock.objects.filter(salespoint=wh, product__is_active=True)
        if kind != 'all':
            qs_stats = qs_stats.filter(product__product_type=kind)
        qs_stats = qs_stats.annotate(remaining=F('opening_qty') + F('transfer_in') - F('transfer_out') - F('sold_qty'))
        total_in_stock = qs_stats.filter(remaining__gt=0).count()
        total_products_all = qs_stats.count()
        low_count = qs_stats.filter(remaining__gt=0, remaining__lte=F('alert_qty')).count()

    # Sort by remaining asc
    items.sort(key=lambda x: (x['remaining_qty'], x['product'].name))

    # Pagination for items
    try:
        page_num = int(request.GET.get('page') or 1)
    except Exception:
        page_num = 1
    try:
        per_page = int(request.GET.get('per_page') or 50)
        if per_page not in [25, 50, 100, 200]:
            per_page = 50
    except Exception:
        per_page = 50
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    paginator = Paginator(items, per_page)
    try:
        page = paginator.page(page_num)
    except PageNotAnInteger:
        page = paginator.page(1)
    except EmptyPage:
        page = paginator.page(paginator.num_pages)

    # Preload existing draft lines by this user (if any)
    draft = WarehousePurchaseRequest.objects.filter(status='draft', requested_by=request.user).order_by('-created_at').first()
    draft_lines = []
    if draft:
        for ln in draft.lines.select_related('product','product__brand'):
            draft_lines.append({
                'product': ln.product,
                'remaining_qty': 0,
                'alert_qty': 0,
                'suggested_qty': int(getattr(ln, 'quantity_requested', 1) or 1),
                'from_draft': True,
            })
    # Merge draft lines (avoid duplicates)
    existing_ids = {r['product'].id for r in items}
    for dl in draft_lines:
        if dl['product'].id not in existing_ids:
            items.append(dl)

    return render(request, 'inventory/warehouse/warehouse_purchase.html', {
        'kind': kind,
        'q': q,
        'rows': page.object_list,
        'paginator': paginator,
        'page': page,
        'per_page': per_page,
        'total_in_stock': total_in_stock,
        'total_products_all': total_products_all,
        'low_count': low_count,
    })


@login_required
def api_wh_cmd_save(request):
    """Save or update a draft WarehousePurchaseRequest for the current user.
    Body: { lines: [{product_id, qty}] }
    """
    role = getattr(request.user, 'role', '')
    if not (request.user.is_superuser or role == 'warehouse_mgr' or getattr(request.user, 'is_staff', False)):
        return JsonResponse({'ok': False, 'error': 'Accès refusé.'}, status=403)
    try:
        payload = json.loads(request.body.decode('utf-8'))
        lines = payload.get('lines') or []
        mode = (payload.get('mode') or '').strip()
        kind = (payload.get('kind') or '').strip() or 'all'
        q = (payload.get('q') or '').strip()
    except Exception:
        return JsonResponse({'ok': False, 'error': 'Requête invalide.'}, status=400)
    if not isinstance(lines, list):
        return JsonResponse({'ok': False, 'error': 'Lignes invalides.'}, status=400)

    # Prepare current user's draft up-front so we can exclude it from active set
    draft_req = WarehousePurchaseRequest.objects.filter(status='draft', requested_by=request.user).order_by('-created_at').first()
    if not draft_req:
        draft_req = WarehousePurchaseRequest.objects.create(requested_by=request.user, status='draft', reference='')

    # If saving all, compute the full low/zero list matching filters and merge
    if mode == 'all':
        wh = SalesPoint.objects.filter(is_warehouse=True).first()
        # Exclude products already in other active purchase requests (sent/acknowledged) and other users' drafts
        active_qs = WarehousePurchaseLine.objects.filter(
            request__status__in=['draft','sent','acknowledged']
        )
        # Exclude current user's active draft from this exclusion to allow re-saving all items
        if draft_req:
            active_qs = active_qs.exclude(request_id=draft_req.id)
        active_product_ids = set(active_qs.values_list('product_id', flat=True))
        products_qs = Product.objects.filter(is_active=True)
        if kind in {'piece','moto'}:
            products_qs = products_qs.filter(product_type=kind)
        if q:
            products_qs = products_qs.filter(Q(name__icontains=q) | Q(brand__name__icontains=q))

        computed_lines = []
        for p in products_qs.select_related('brand'):
            if p.id in active_product_ids:
                continue
            sps = None
            if wh:
                try:
                    sps = SalesPointStock.objects.get(salespoint=wh, product=p)
                except SalesPointStock.DoesNotExist:
                    sps = None
            remaining = int(getattr(sps, 'remaining_qty', 0) or 0)
            alert = int(getattr(sps, 'alert_qty', 5) or 5)
            if remaining <= alert:
                computed_lines.append({'product_id': p.id, 'qty': 1})

        # Merge provided lines (from UI list) over computed ones, prefer provided qty
        qty_by_pid = {int(l.get('product_id') or 0): int(l.get('qty') or 1) or 1 for l in computed_lines}
        for l in lines:
            pid = int(l.get('product_id') or 0)
            if pid <= 0:
                continue
            qty_by_pid[pid] = int(l.get('qty') or 1) or 1
        # Rebuild lines
        lines = [{'product_id': pid, 'qty': qty_by_pid[pid]} for pid in qty_by_pid.keys() if pid > 0]

    with transaction.atomic():
        req = draft_req
        # Reset lines
        req.lines.all().delete()
        created = 0
        for ln in lines:
            pid = int(ln.get('product_id') or 0)
            qty = int(ln.get('qty') or 1) or 1
            if pid <= 0:
                continue
            WarehousePurchaseLine.objects.create(request=req, product_id=pid, quantity_requested=qty)
            created += 1

    # Notify current warehouse manager about draft save/update
    try:
        from apps.sales.models import Notification
        Notification.objects.create(
            user=request.user,
            message=f"📝 Brouillon CMD-WH enregistré ({created} article(s))",
            link="/inventory/warehouse/purchase/",
            kind="cmd_wh_draft",
        )
    except Exception:
        pass

    return JsonResponse({'ok': True, 'draft_id': req.id, 'count': created})


@login_required
def api_wh_cmd_search_products(request):
    """Search products to add to the warehouse purchase list (even if not low).
    Query params: q, type in {'piece','moto','all'}
    Excludes products already present in active purchase requests (draft/sent/acknowledged).
    """
    role = getattr(request.user, 'role', '')
    if not (request.user.is_superuser or role == 'warehouse_mgr' or getattr(request.user, 'is_staff', False)):
        return JsonResponse([], safe=False)
    q = (request.GET.get('q') or '').strip()
    kind = (request.GET.get('type') or 'all').strip()
    wh = SalesPoint.objects.filter(is_warehouse=True).first()
    # Active product ids to exclude
    active_ids = set(
        WarehousePurchaseLine.objects.filter(request__status__in=['draft','sent','acknowledged']).values_list('product_id', flat=True)
    )
    prods = Product.objects.filter(is_active=True)
    if kind in {'piece','moto'}:
        prods = prods.filter(product_type=kind)
    if q:
        prods = prods.filter(Q(name__icontains=q) | Q(brand__name__icontains=q))
    rows = []
    for p in prods.select_related('brand')[:120]:
        if p.id in active_ids:
            continue
        sps = None
        if wh:
            try:
                sps = SalesPointStock.objects.get(salespoint=wh, product=p)
            except SalesPointStock.DoesNotExist:
                sps = None
        remaining = int(getattr(sps, 'remaining_qty', 0) or 0)
        alert = int(getattr(sps, 'alert_qty', 5) or 5)
        threshold = alert
        suggested = max(1, alert * 3 - remaining)
        # Only show products that (1) have positive stock and (2) were not auto-listed
        # i.e., exclude finishing products already shown on the main list (remaining <= threshold)
        if remaining <= threshold:
            continue
        rows.append({
            'product_id': p.id,
            'name': p.name,
            'brand': getattr(getattr(p, 'brand', None), 'name', '') or '',
            'remaining_qty': remaining,
            'alert_qty': alert,
            'suggested_qty': suggested,
        })
    return JsonResponse({'ok': True, 'rows': rows})


@login_required
def warehouse_dashboard(request):
    if not (request.user.is_superuser or getattr(request.user, 'role', '') == 'warehouse_mgr' or getattr(request.user, 'is_staff', False)):
        return redirect('sales:dashboard')
    # Basic metrics - only count actual commandes (salespoint requests), not warehouse restocks
    pending = RestockRequest.objects.filter(status='sent', reference__startswith='WH-RQ-').count()
    confirmed_restocks = RestockRequest.objects.filter(status__in=['validated', 'partially_validated'], reference__startswith='WH-RQ-').count()
    total_confirmed_qty = RestockRequest.objects.filter(
        status__in=['validated', 'partially_validated'],
        reference__startswith='WH-RQ-'
    ).aggregate(
        total=Sum('lines__quantity_approved')
    )['total'] or 0
    recent_tx = StockTransaction.objects.order_by('-created_at')[:10]

    # Filters (type tabs, search, pagination)
    kind = (request.GET.get('type') or 'piece').strip()  # 'piece' | 'moto'
    q = (request.GET.get('q') or '').strip()
    try:
        per_page = int(request.GET.get('pp') or 50)
    except Exception:
        per_page = 50
    per_page = 50 if per_page not in (25, 50, 100, 200) else per_page
    try:
        page_num = int(request.GET.get('page') or 1)
    except Exception:
        page_num = 1

    # Try to target the warehouse salespoint heuristically
    wh = SalesPoint.objects.filter(is_warehouse=True).first()
    if not wh:
        wh = SalesPoint.objects.filter(Q(name__icontains='entrep') | Q(name__icontains='ware')).order_by('name').first()
    base = SalesPointStock.objects.select_related('product', 'product__brand')
    if wh:
        base = base.filter(salespoint=wh)
    # else: show across all points as fallback

    qs = base.filter(product__is_active=True, product__product_type=kind).order_by('product__name')
    if q:
        qs = qs.filter(Q(product__name__icontains=q) | Q(product__brand__name__icontains=q))

    # Add transit quantity calculation for each product
    for stock in qs:
        # Calculate transit quantity (sent but not yet validated)
        transit_qty = RestockLine.objects.filter(
            request__status__in=['sent', 'partially_validated'],
            product=stock.product,
            request__salespoint__is_warehouse=False  # Only outgoing from warehouse
        ).aggregate(
            total=Sum('quantity_approved')
        )['total'] or 0
        
        stock.transit_qty = transit_qty
        
        # Calculate confirmed restock quantity (sold from warehouse to salespoints and confirmed as received)
        # This represents "Qté vendue" - quantity sold/restocked from warehouse
        confirmed_qty = RestockLine.objects.filter(
            request__status__in=['validated', 'partially_validated'],
            product=stock.product,
            request__salespoint__is_warehouse=False,  # Only outgoing from warehouse to salespoints
            validated_at__isnull=False  # Only validated lines
        ).aggregate(
            total=Sum('quantity_approved')
        )['total'] or 0
        
        stock.confirmed_qty = confirmed_qty

    # Paginate
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    paginator = Paginator(qs, per_page)
    try:
        page = paginator.page(page_num)
    except PageNotAnInteger:
        page = paginator.page(1)
    except EmptyPage:
        page = paginator.page(paginator.num_pages)

    return render(
        request,
        'inventory/warehouse/warehouse_dashboard.html',
        {
            'pending': pending,
            'confirmed_restocks': confirmed_restocks,
            'total_confirmed_qty': total_confirmed_qty,
            'recent_tx': recent_tx,
            'warehouse_rows': page.object_list,
            'page': page,
            'paginator': paginator,
            'q': q,
            'kind': kind,
            'per_page': per_page,
        },
    )


@login_required
def warehouse_requests(request):
    """List restock requests for the warehouse manager/staff.
    Filters: q (salespoint or product), status (sent/approved/rejected/fulfilled), date range.
    """
    # Allow warehouse managers, superusers, and staff
    if not (request.user.is_superuser or getattr(request.user, 'role', '') == 'warehouse_mgr' or getattr(request.user, 'is_staff', False)):
        return redirect('sales:dashboard')

    q = (request.GET.get('q') or '').strip()
    status = (request.GET.get('status') or 'sent').strip()  # Default to 'sent' to show incoming requests
    df = request.GET.get('from') or ''
    dt = request.GET.get('to') or ''

    qs = RestockRequest.objects.select_related('salespoint', 'requested_by').order_by('-created_at')
    # Only show salespoint-initiated requests (WH-RQ-...)
    qs = qs.filter(reference__startswith='WH-RQ-')
    if status in {'draft', 'sent', 'approved', 'rejected', 'fulfilled', 'cancelled', 'partially_validated', 'validated'}:
        qs = qs.filter(status=status)
    if q:
        qs = qs.filter(
            Q(salespoint__name__icontains=q) |
            Q(requested_by__username__icontains=q) |
            Q(lines__product__name__icontains=q) |
            Q(lines__product__brand__name__icontains=q)
        ).distinct()
    if df:
        qs = qs.filter(created_at__date__gte=df)
    if dt:
        qs = qs.filter(created_at__date__lte=dt)

    return render(request, 'inventory/warehouse/warehouse_requests.html', {
        'rows': qs[:500],
        'q': q,
        'status': status,
        'date_from': df,
        'date_to': dt,
    })


@login_required
def warehouse_request_lines(request, req_id: int):
    """Return JSON of lines for a restock request (warehouse view).
    Optimized to avoid massive IN() prefetch queries on large requests.
    """
    if not (request.user.is_superuser or getattr(request.user, 'role', '') == 'warehouse_mgr' or getattr(request.user, 'is_staff', False)):
        return JsonResponse({"ok": False, "error": "Accès refusé."}, status=403)

    # Load header without prefetching huge related sets
    req = get_object_or_404(
        RestockRequest.objects.select_related('salespoint', 'requested_by'),
        pk=req_id,
    )

    # Stream lines with joins (avoids big IN clause prefetch)
    lines_qs = (
        RestockLine.objects
        .select_related('product', 'product__brand')
        .filter(request_id=req.id)
        .only('product_id', 'remaining_qty', 'quantity_requested', 'quantity_approved', 'validated_at', 'product__name', 'product__brand__name')
        .order_by('product__name')
    )

    data = {
        "ok": True,
        "id": req.id,
        "reference": req.reference or f"#{req.id}",
        "salespoint": getattr(req.salespoint, 'name', ''),
        "sent_by": getattr(req.requested_by, 'username', ''),
        "status": req.status,
        "created_at": req.created_at.strftime('%d/%m/%Y %H:%M') if req.created_at else '',
        "lines": [
            {
                "product_id": ln.product_id,
                "name": getattr(ln.product, 'name', f"#{ln.product_id}"),
                "brand": getattr(getattr(ln.product, 'brand', None), 'name', ''),
                # Show the quantity chosen by the warehouse manager (approved → legacy quantity → requested)
                "qty_sent": int((ln.quantity_approved or ln.quantity or ln.quantity_requested or 0) or 0),
                # Add validation status
                "is_validated": ln.validated_at is not None,
                "validated_at": ln.validated_at.strftime('%d/%m/%Y %H:%M') if ln.validated_at else None,
            }
            for ln in lines_qs
        ],
    }
    return JsonResponse(data)


@login_required
def warehouse_journal(request):
    """Warehouse journal - same as restock journal for consistency."""
    if not (request.user.is_superuser or getattr(request.user, 'role', '') == 'warehouse_mgr' or getattr(request.user, 'is_staff', False)):
        return redirect('sales:dashboard')
    
    q = (request.GET.get('q') or '').strip()
    status_filter = (request.GET.get('status') or '').strip()
    sp_id = int(request.GET.get('sp') or 0)
    date_filter = (request.GET.get('date') or '').strip()
    start_date = request.GET.get('start_date') or ''
    end_date = request.GET.get('end_date') or ''
    
    # Get all restock requests from warehouse
    qs = RestockRequest.objects.select_related('salespoint').prefetch_related('lines__product').order_by('-created_at')
    # Exclude salespoint requests (WH-RQ-...) and inbound from Commercial Director (CD-...)
    qs = qs.exclude(reference__startswith='WH-RQ-').exclude(reference__startswith='CD-')
    
    # Apply filters
    if status_filter == 'not_validated':
        # Show requests that are not fully validated (sent or partially validated)
        qs = qs.filter(status__in=['sent', 'partially_validated'])
    elif status_filter in ['sent', 'partially_validated', 'validated']:
        qs = qs.filter(status=status_filter)
    
    if sp_id:
        qs = qs.filter(salespoint_id=sp_id)
    
    # Apply date filters
    today = timezone.localdate()
    if date_filter == 'today':
        qs = qs.filter(created_at__date=today)
    elif date_filter == 'yesterday':
        qs = qs.filter(created_at__date=today - timedelta(days=1))
    elif date_filter == 'week':
        week_start = today - timedelta(days=today.weekday())
        qs = qs.filter(created_at__date__gte=week_start)
    elif date_filter == 'custom':
        # Custom date range
        if start_date:
            try:
                start_dt = timezone.datetime.strptime(start_date, '%Y-%m-%d').date()
                qs = qs.filter(created_at__date__gte=start_dt)
            except ValueError:
                pass
        if end_date:
            try:
                end_dt = timezone.datetime.strptime(end_date, '%Y-%m-%d').date()
                qs = qs.filter(created_at__date__lte=end_dt)
            except ValueError:
                pass
    elif not date_filter and not q and not status_filter and not sp_id:
        # Default: show today's restocks if no other filters are applied
        qs = qs.filter(created_at__date=today)
    
    if q:
        qs = qs.filter(
            Q(reference__icontains=q) | 
            Q(salespoint__name__icontains=q) | 
            Q(lines__product__name__icontains=q)
        ).distinct()
    
    # Convert to list to avoid queryset evaluation issues
    rows = list(qs[:100])
    
    # Calculate totals for each restock request
    for req in rows:
        lines = list(req.lines.all())  # Convert to list to avoid queryset issues
        
        req.total_products = len(lines)
        req.calculated_total_quantity = sum(
            (line.quantity_approved or line.quantity_requested or 0) 
            for line in lines
        )
        req.total_value = sum(
            (line.quantity_approved or line.quantity_requested or 0) * (line.product.cost_price or 0)
            for line in lines
        )
        
        validated_lines = [line for line in lines if line.validated_at is not None]
        req.validated_products = len(validated_lines)
        req.validated_quantity = sum(
            (line.quantity_approved or line.quantity_requested or 0) 
            for line in validated_lines
        )
        req.validated_value = sum(
            (line.quantity_approved or line.quantity_requested or 0) * (line.product.cost_price or 0)
            for line in validated_lines
        )
    
    # Calculate summary statistics
    total_requests = len(rows)
    pending_requests = len([req for req in rows if req.status == 'sent'])
    validated_requests = len([req for req in rows if req.status == 'validated'])
    total_value = sum(req.total_value for req in rows)
    
    return render(request, 'inventory/warehouse/warehouse_journal.html', {
        'rows': rows,
        'total_requests': total_requests,
        'pending_requests': pending_requests,
        'validated_requests': validated_requests,
        'total_value': total_value,
        'q': q,
        'status_filter': status_filter,
        'date_filter': date_filter,
        'start_date': start_date,
        'end_date': end_date,
        'salespoints': SalesPoint.objects.filter(is_warehouse=False).order_by('name'),
        'sp_id': sp_id,
    })


@login_required
def api_wh_restock_lines(request, req_id: int):
    """Return JSON lines for a RestockRequest for use by the inbound modal."""
    if not (request.user.is_superuser or getattr(request.user, 'role', '') in ['warehouse_mgr'] or getattr(request.user, 'is_staff', False)):
        return JsonResponse({'ok': False, 'error': 'Accès refusé.'}, status=403)
    req = get_object_or_404(RestockRequest.objects.select_related('provider'), pk=req_id)
    lines = []
    for ln in req.lines.select_related('product','product__brand').all():
        qty = int(ln.quantity or ln.quantity_requested or 0)
        lines.append({
            'product_id': ln.product_id,
            'product': f"{ln.product.name} {('• ' + ln.product.brand.name) if ln.product.brand_id else ''}",
            'qty': qty,
        })
    return JsonResponse({'ok': True, 'reference': req.reference or '', 'provider': getattr(req.provider, 'name', ''), 'lines': lines})


@login_required
def salespoints_stock(request):
    """Enhanced warehouse stocks view showing both warehouse and salespoint stocks."""
    if not (request.user.is_superuser or getattr(request.user, 'role', '') == 'warehouse_mgr' or getattr(request.user, 'is_staff', False)):
        return redirect('sales:dashboard')
    
    # Filters
    sp_id = int(request.GET.get('sp') or 0)
    q = (request.GET.get('q') or '').strip()
    view_type = request.GET.get('view', 'warehouse')  # 'warehouse' or 'salespoint'
    product_type = request.GET.get('type', 'piece')  # 'piece' or 'moto'
    stock_filter = (request.GET.get('stock') or 'all').strip()  # 'all' | 'low' | 'zero' | 'ok'
    
    # Get all salespoints
    sps = SalesPoint.objects.order_by('name')
    
    # Try to identify warehouse salespoint
    warehouse_sp = SalesPoint.objects.filter(is_warehouse=True).first()
    if not warehouse_sp:
        warehouse_sp = SalesPoint.objects.filter(Q(name__icontains='entrep') | Q(name__icontains='ware')).order_by('name').first()
    
    # Get products with stock data
    products = Product.objects.filter(is_active=True, product_type=product_type)
    if q:
        products = products.filter(Q(name__icontains=q) | Q(brand__name__icontains=q))
    
    # Prepare data based on view type
    data = []
    selected_salespoint = None
    
    # Pagination settings
    try:
        page_num = int(request.GET.get('page', 1))
    except (ValueError, TypeError):
        page_num = 1
    
    try:
        per_page = int(request.GET.get('per_page', 50))
        # Limit per_page to reasonable values
        if per_page not in [25, 50, 100, 200]:
            per_page = 50
    except (ValueError, TypeError):
        per_page = 50
    
    if view_type == 'warehouse':
        # Show ALL products - both with and without warehouse stock
        for product in products:
            wh_stock = None
            if warehouse_sp:
                try:
                    wh_stock = SalesPointStock.objects.get(salespoint=warehouse_sp, product=product)
                except SalesPointStock.DoesNotExist:
                    wh_stock = None
            
            # Calculate transit quantity (sent but not yet validated)
            transit_qty = RestockLine.objects.filter(
                request__status__in=['sent', 'partially_validated'],
                product=product,
                request__salespoint__is_warehouse=False
            ).aggregate(total=Sum('quantity_approved'))['total'] or 0
            
            # Calculate confirmed restock quantity (sold from warehouse to salespoints)
            confirmed_qty = RestockLine.objects.filter(
                request__status__in=['validated', 'partially_validated'],
                product=product,
                request__salespoint__is_warehouse=False,
                validated_at__isnull=False
            ).aggregate(total=Sum('quantity_approved'))['total'] or 0
            
            # Add product with stock data (or zeros if no stock)
            data.append({
                'product': product,
                'has_stock': wh_stock is not None,
                'opening_qty': wh_stock.opening_qty if wh_stock else 0,
                'sold_qty': wh_stock.sold_qty if wh_stock else 0,
                'remaining_qty': wh_stock.remaining_qty if wh_stock else 0,
                'available_qty': wh_stock.available_qty if wh_stock else 0,
                'reserved_qty': wh_stock.reserved_qty if wh_stock else 0,
                'transit_qty': transit_qty,
                'confirmed_qty': confirmed_qty,
            })
        
        # Optional filtering by stock status
        if stock_filter in {'low','zero','ok'}:
            def _wh_match(item):
                rem = int(item.get('remaining_qty') or 0)
                if stock_filter == 'zero':
                    return rem == 0
                if stock_filter == 'low':
                    return rem > 0 and rem <= 5
                return rem > 5
            data = [it for it in data if _wh_match(it)]

        # Sort by stock status first (products with stock first), then by name
        data.sort(key=lambda x: (not x['has_stock'], x['product'].name))
        
    elif view_type == 'salespoint' and sp_id:
        # Show ALL products for selected salespoint - both with and without stock
        try:
            selected_salespoint = SalesPoint.objects.get(id=sp_id)
        except SalesPoint.DoesNotExist:
            selected_salespoint = None
        
        if selected_salespoint:
            for product in products:
                sp_stock = None
                try:
                    sp_stock = SalesPointStock.objects.get(salespoint=selected_salespoint, product=product)
                except SalesPointStock.DoesNotExist:
                    sp_stock = None
                
                # Add product with stock data (or zeros if no stock)
                data.append({
                    'product': product,
                    'has_stock': sp_stock is not None,
                    'opening_qty': sp_stock.opening_qty if sp_stock else 0,
                    'sold_qty': sp_stock.sold_qty if sp_stock else 0,
                    'remaining_qty': sp_stock.remaining_qty if sp_stock else 0,
                    'available_qty': sp_stock.available_qty if sp_stock else 0,
                    'reserved_qty': sp_stock.reserved_qty if sp_stock else 0,
                    'alert_qty': sp_stock.alert_qty if sp_stock else 5,  # Default alert quantity
                })
            
            # Optional filtering by stock status
            if stock_filter in {'low','zero','ok'}:
                def _sp_match(item):
                    rem = int(item.get('remaining_qty') or 0)
                    alert = int(item.get('alert_qty') or 5)
                    if stock_filter == 'zero':
                        return rem == 0
                    if stock_filter == 'low':
                        return rem > 0 and rem <= alert
                    return rem > alert
                data = [it for it in data if _sp_match(it)]

            # Sort by stock status first (products with stock first), then by name
            data.sort(key=lambda x: (not x['has_stock'], x['product'].name))
    
    # Apply pagination
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    paginator = Paginator(data, per_page)
    
    try:
        page = paginator.page(page_num)
    except PageNotAnInteger:
        page = paginator.page(1)
    except EmptyPage:
        page = paginator.page(paginator.num_pages)
    
    # Calculate statistics for the current page
    products_with_stock = sum(1 for item in page.object_list if item['has_stock'])
    products_without_stock = sum(1 for item in page.object_list if not item['has_stock'])
    
    return render(request, 'inventory/warehouse/salespoints_stock.html', {
        'salespoints': sps,
        'sp_id': sp_id,
        'q': q,
        'view_type': view_type,
        'product_type': product_type,
        'stock_filter': stock_filter,
        'warehouse_sp': warehouse_sp,
        'selected_salespoint': selected_salespoint,
        'data': page.object_list,
        'page': page,
        'paginator': paginator,
        'products_with_stock': products_with_stock,
        'products_without_stock': products_without_stock,
    })


@login_required
def export_finished_products(request):
    """Export products that are finished or getting finished to Excel."""
    if not (request.user.is_superuser or getattr(request.user, 'role', '') == 'warehouse_mgr' or getattr(request.user, 'is_staff', False)):
        return redirect('sales:dashboard')
    
    # Get filters from request
    view_type = request.GET.get('view', 'warehouse')
    product_type = request.GET.get('type', 'piece')
    sp_id = int(request.GET.get('sp') or 0)
    q = (request.GET.get('q') or '').strip()
    
    # Get all salespoints
    sps = SalesPoint.objects.order_by('name')
    
    # Try to identify warehouse salespoint
    warehouse_sp = SalesPoint.objects.filter(is_warehouse=True).first()
    if not warehouse_sp:
        warehouse_sp = SalesPoint.objects.filter(Q(name__icontains='entrep') | Q(name__icontains='ware')).order_by('name').first()
    
    # Get products with stock data
    products = Product.objects.filter(is_active=True, product_type=product_type)
    if q:
        products = products.filter(Q(name__icontains=q) | Q(brand__name__icontains=q))
    
    # Prepare data based on view type
    data = []
    selected_salespoint = None
    
    if view_type == 'warehouse':
        # Show ALL products - both with and without warehouse stock
        for product in products:
            wh_stock = None
            if warehouse_sp:
                try:
                    wh_stock = SalesPointStock.objects.get(salespoint=warehouse_sp, product=product)
                except SalesPointStock.DoesNotExist:
                    wh_stock = None
            
            # Calculate transit quantity (sent but not yet validated)
            transit_qty = RestockLine.objects.filter(
                request__status__in=['sent', 'partially_validated'],
                product=product,
                request__salespoint__is_warehouse=False
            ).aggregate(total=Sum('quantity_approved'))['total'] or 0
            
            # Calculate confirmed restock quantity (sold from warehouse to salespoints)
            confirmed_qty = RestockLine.objects.filter(
                request__status__in=['validated', 'partially_validated'],
                product=product,
                request__salespoint__is_warehouse=False,
                validated_at__isnull=False
            ).aggregate(total=Sum('quantity_approved'))['total'] or 0
            
            # Add product with stock data (or zeros if no stock)
            data.append({
                'product': product,
                'has_stock': wh_stock is not None,
                'opening_qty': wh_stock.opening_qty if wh_stock else 0,
                'sold_qty': wh_stock.sold_qty if wh_stock else 0,
                'remaining_qty': wh_stock.remaining_qty if wh_stock else 0,
                'available_qty': wh_stock.available_qty if wh_stock else 0,
                'reserved_qty': wh_stock.reserved_qty if wh_stock else 0,
                'transit_qty': transit_qty,
                'confirmed_qty': confirmed_qty,
            })
        
    elif view_type == 'salespoint' and sp_id:
        # Show ALL products for selected salespoint - both with and without stock
        try:
            selected_salespoint = SalesPoint.objects.get(id=sp_id)
        except SalesPoint.DoesNotExist:
            selected_salespoint = None
        
        if selected_salespoint:
            for product in products:
                sp_stock = None
                try:
                    sp_stock = SalesPointStock.objects.get(salespoint=selected_salespoint, product=product)
                except SalesPointStock.DoesNotExist:
                    sp_stock = None
                
                # Add product with stock data (or zeros if no stock)
                data.append({
                    'product': product,
                    'has_stock': sp_stock is not None,
                    'opening_qty': sp_stock.opening_qty if sp_stock else 0,
                    'sold_qty': sp_stock.sold_qty if sp_stock else 0,
                    'remaining_qty': sp_stock.remaining_qty if sp_stock else 0,
                    'available_qty': sp_stock.available_qty if sp_stock else 0,
                    'reserved_qty': sp_stock.reserved_qty if sp_stock else 0,
                    'alert_qty': sp_stock.alert_qty if sp_stock else 5,  # Default alert quantity
                })
    
    # Filter for finished or getting finished products
    finished_products = []
    for item in data:
        if not item['has_stock'] or item['remaining_qty'] <= 5:  # No stock or low stock
            finished_products.append(item)
    
    # Sort by stock status first (products with stock first), then by name
    finished_products.sort(key=lambda x: (not x['has_stock'], x['product'].name))
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Produits Finis - En Finition"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    if view_type == 'warehouse':
        headers = [
            'ID Produit', 'Nom du Produit', 'Marque', 'Type', 'Statut Stock',
            'Ouverture', 'Vendu', 'Restant', 'Disponible', 'Réservé',
            'En Transit', 'Confirmé', 'Prix Achat', 'Prix Vente'
        ]
    else:
        headers = [
            'ID Produit', 'Nom du Produit', 'Marque', 'Type', 'Statut Stock',
            'Ouverture', 'Vendu', 'Restant', 'Réservé', 'Disponible',
            'Seuil', 'Prix Achat', 'Prix Vente'
        ]
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Write data
    for row, item in enumerate(finished_products, 2):
        # Determine stock status
        if not item['has_stock']:
            status = "Sans stock"
        elif item['remaining_qty'] == 0:
            status = "Épuisé"
        elif item['remaining_qty'] <= 5:
            status = "Stock bas"
        else:
            status = "En stock"
        
        # Basic product info
        ws.cell(row=row, column=1, value=item['product'].id)
        ws.cell(row=row, column=2, value=item['product'].name)
        ws.cell(row=row, column=3, value=item['product'].brand.name if item['product'].brand else '')
        ws.cell(row=row, column=4, value=item['product'].product_type)
        ws.cell(row=row, column=5, value=status)
        
        # Stock quantities
        ws.cell(row=row, column=6, value=item['opening_qty'])
        ws.cell(row=row, column=7, value=item['sold_qty'])
        ws.cell(row=row, column=8, value=item['remaining_qty'])
        
        if view_type == 'warehouse':
            ws.cell(row=row, column=9, value=item['available_qty'])
            ws.cell(row=row, column=10, value=item['reserved_qty'])
            ws.cell(row=row, column=11, value=item['transit_qty'])
            ws.cell(row=row, column=12, value=item['confirmed_qty'])
            ws.cell(row=row, column=13, value=float(item['product'].cost_price) if item['product'].cost_price else 0)
            ws.cell(row=row, column=14, value=float(item['product'].selling_price) if item['product'].selling_price else 0)
        else:
            ws.cell(row=row, column=9, value=item['reserved_qty'])
            ws.cell(row=row, column=10, value=item['available_qty'])
            ws.cell(row=row, column=11, value=item['alert_qty'])
            ws.cell(row=row, column=12, value=float(item['product'].cost_price) if item['product'].cost_price else 0)
            ws.cell(row=row, column=13, value=float(item['product'].selling_price) if item['product'].selling_price else 0)
        
        # Apply borders to all cells
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).border = border
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Add summary information
    summary_row = len(finished_products) + 3
    ws.cell(row=summary_row, column=1, value="RÉSUMÉ").font = Font(bold=True, size=14)
    ws.cell(row=summary_row + 1, column=1, value=f"Total produits finis/en finition: {len(finished_products)}")
    ws.cell(row=summary_row + 2, column=1, value=f"Vue: {'Entrepôt' if view_type == 'warehouse' else 'Point de vente'}")
    if selected_salespoint:
        ws.cell(row=summary_row + 3, column=1, value=f"Point de vente: {selected_salespoint.name}")
    ws.cell(row=summary_row + 4, column=1, value=f"Type de produit: {product_type}")
    ws.cell(row=summary_row + 5, column=1, value=f"Date d'export: {timezone.now().strftime('%d/%m/%Y %H:%M')}")
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Generate filename
    location = "entrepot" if view_type == 'warehouse' else f"point_vente_{selected_salespoint.name.lower().replace(' ', '_')}" if selected_salespoint else "point_vente"
    filename = f"produits_finis_{location}_{product_type}_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Save workbook to response
    wb.save(response)
    return response


@login_required
def restock_history(request):
    if not (request.user.is_superuser or getattr(request.user, 'role', '') == 'warehouse_mgr' or getattr(request.user, 'is_staff', False)):
        return redirect('sales:dashboard')
    sp_id = int(request.GET.get('sp') or 0)
    product_q = (request.GET.get('q') or '').strip()
    reqs = RestockRequest.objects.select_related('salespoint').order_by('-created_at')
    if sp_id:
        reqs = reqs.filter(salespoint_id=sp_id)
    if product_q:
        reqs = reqs.filter(lines__product__name__icontains=product_q).distinct()
    return render(request, 'inventory/reports/restock_history.html', { 'rows': reqs[:500], 'salespoints': SalesPoint.objects.all(), 'sp_id': sp_id, 'q': product_q })


@login_required
def warehouse_restock_journal(request):
    """Warehouse journal specifically for restock requests with detailed status tracking."""
    if not (request.user.is_superuser or getattr(request.user, 'role', '') == 'warehouse_mgr' or getattr(request.user, 'is_staff', False)):
        return redirect('sales:dashboard')
    
    q = (request.GET.get('q') or '').strip()
    status_filter = (request.GET.get('status') or '').strip()
    sp_id = int(request.GET.get('sp') or 0)
    date_filter = (request.GET.get('date') or '').strip()
    start_date = request.GET.get('start_date') or ''
    end_date = request.GET.get('end_date') or ''
    
    # Debug: Print filter parameters
    print(f"DEBUG: Filter parameters - q: '{q}', status: '{status_filter}', sp_id: {sp_id}, date: '{date_filter}'")
    print(f"DEBUG: Date range - start: '{start_date}', end: '{end_date}'")
    print(f"DEBUG: All GET parameters: {dict(request.GET)}")
    
    # Get all restock requests from warehouse
    qs = RestockRequest.objects.select_related('salespoint').prefetch_related('lines__product').order_by('-created_at')
    
    # Apply filters
    if status_filter == 'not_validated':
        # Show requests that are not fully validated (sent or partially validated)
        qs = qs.filter(status__in=['sent', 'partially_validated'])
    elif status_filter in ['sent', 'partially_validated', 'validated']:
        qs = qs.filter(status=status_filter)
    
    if sp_id:
        qs = qs.filter(salespoint_id=sp_id)
    
    # Apply date filters
    today = timezone.localdate()
    if date_filter == 'today':
        qs = qs.filter(created_at__date=today)
    elif date_filter == 'yesterday':
        qs = qs.filter(created_at__date=today - timedelta(days=1))
    elif date_filter == 'week':
        week_start = today - timedelta(days=today.weekday())
        qs = qs.filter(created_at__date__gte=week_start)
    elif date_filter == 'custom':
        # Custom date range
        if start_date:
            try:
                start_dt = timezone.datetime.strptime(start_date, '%Y-%m-%d').date()
                qs = qs.filter(created_at__date__gte=start_dt)
            except ValueError:
                pass
        if end_date:
            try:
                end_dt = timezone.datetime.strptime(end_date, '%Y-%m-%d').date()
                qs = qs.filter(created_at__date__lte=end_dt)
            except ValueError:
                pass
    elif not date_filter and not q and not status_filter and not sp_id:
        # Default: show today's restocks if no other filters are applied
        qs = qs.filter(created_at__date=today)
    
    if q:
        qs = qs.filter(
            Q(reference__icontains=q) | 
            Q(salespoint__name__icontains=q) | 
            Q(lines__product__name__icontains=q)
        ).distinct()
    
    # Convert to list to avoid queryset evaluation issues
    rows = list(qs[:100])
    
    # Calculate totals for each restock request
    for req in rows:
        lines = list(req.lines.all())  # Convert to list to avoid queryset issues
        
        req.total_products = len(lines)
        req.calculated_total_quantity = sum(
            (line.quantity_approved or line.quantity_requested or 0) 
            for line in lines
        )
        req.total_value = sum(
            (line.quantity_approved or line.quantity_requested or 0) * (line.product.cost_price or 0)
            for line in lines
        )
        
        validated_lines = [line for line in lines if line.validated_at is not None]
        req.validated_products = len(validated_lines)
        req.validated_quantity = sum(
            (line.quantity_approved or line.quantity_requested or 0) 
            for line in validated_lines
        )
        req.validated_value = sum(
            (line.quantity_approved or line.quantity_requested or 0) * (line.product.cost_price or 0)
            for line in validated_lines
        )
    
    return render(request, 'inventory/warehouse/warehouse_journal.html', {
        'rows': rows,
        'q': q,
        'status_filter': status_filter,
        'date_filter': date_filter,
        'start_date': start_date,
        'end_date': end_date,
        'salespoints': SalesPoint.objects.filter(is_warehouse=False).order_by('name'),
        'sp_id': sp_id,
    })


# ===== Warehouse restock APIs =====

@login_required
def api_wh_stock(request):
    """Search warehouse stock by type and query for the restock modal."""
    role = getattr(request.user, 'role', '')
    if not (request.user.is_superuser or role == 'warehouse_mgr' or getattr(request.user, 'is_staff', False)):
        return JsonResponse([], safe=False)
    wh = SalesPoint.objects.filter(is_warehouse=True).first()
    if not wh:
        return JsonResponse([], safe=False)
    kind = (request.GET.get('type') or 'piece').strip()
    q = (request.GET.get('q') or '').strip()
    qs = SalesPointStock.objects.select_related('product', 'product__brand').filter(salespoint=wh, product__is_active=True)
    if kind in {'piece','moto'}:
        qs = qs.filter(product__product_type=kind)
    if q:
        qs = qs.filter(Q(product__name__icontains=q) | Q(product__brand__name__icontains=q))
    rows = []
    for sps in qs.order_by('product__name')[:120]:
        rows.append({
            'product_id': sps.product_id,
            'name': sps.product.name,
            'brand': getattr(getattr(sps.product, 'brand', None), 'name', ''),
            'available': int(getattr(sps, 'available_qty', 0) or 0),
            'price': str(getattr(sps.product, 'wholesale_price', 0) or 0),
        })
    return JsonResponse({'ok': True, 'rows': rows})


@login_required
def api_wh_restock_send(request):
    """Create transfers from warehouse to a destination salespoint.
    Body: { to_sp: int, kind: 'P'|'M', lines: [{product_id, qty}, ...] }
    Generates a daily reference WH-DDMMYY-P-XXXX used across created Transfer rows.
    """
    role = getattr(request.user, 'role', '')
    if not (request.user.is_superuser or role == 'warehouse_mgr' or getattr(request.user, 'is_staff', False)):
        return JsonResponse({'ok': False, 'error': 'Accès refusé.'}, status=403)
    wh = SalesPoint.objects.filter(is_warehouse=True).first()
    if not wh:
        return JsonResponse({'ok': False, 'error': "Entrepôt introuvable."}, status=400)
    try:
        payload = json.loads(request.body.decode('utf-8'))
        to_sp = int(payload.get('to_sp') or 0)
        kind = (payload.get('kind') or 'P').upper()
        lines = payload.get('lines') or []
    except Exception:
        return JsonResponse({'ok': False, 'error': 'Requête invalide.'}, status=400)
    if kind not in {'P','M'}:
        kind = 'P'
    if not to_sp:
        return JsonResponse({'ok': False, 'error': 'Point de vente destination requis.'}, status=400)
    if not isinstance(lines, list) or not lines:
        return JsonResponse({'ok': False, 'error': 'Aucun article.'}, status=400)

    today = timezone.localdate()
    with transaction.atomic():
        # Compute next sequence for today safely (prefix WH)
        prefix = f"WH-{today.strftime('%d%m%y')}-{kind}-"
        # Count existing transfers today using reference prefix
        base = StockTransaction.objects.filter(reference__startswith=prefix).values('reference').annotate(c=Count('id'))
        # Fall back to scanning Transfers if StockTransaction is empty
        try:
            used = 0
            for r in base:
                try:
                    num = int(str(r['reference']).split('-')[-1])
                except Exception:
                    num = 0
                used = max(used, num)
        except Exception:
            used = 0
        seq = used + 1
        ref = f"{prefix}{seq:04d}"

        # Create a RestockRequest instead of individual Transfer objects
        restock_request = RestockRequest.objects.create(
            salespoint_id=to_sp,
            requested_by=request.user,
            status='sent',  # Sent from warehouse, awaiting validation
            reference=ref,
            sent_at=timezone.now(),  # Record when it was sent
        )
        
        created = 0
        for ln in lines:
            pid = int(ln.get('product_id') or 0)
            qty = int(ln.get('qty') or 0)
            if not pid or qty <= 0:
                continue
            # Ensure warehouse has the product row and deduct stock immediately
            sps, created_sps = SalesPointStock.objects.get_or_create(
                salespoint=wh, 
                product_id=pid,
                defaults={'opening_qty': 0, 'sold_qty': 0, 'transfer_in': 0, 'transfer_out': 0, 'alert_qty': 0, 'reserved_qty': 0}
            )
            
            available = int(getattr(sps, 'available_qty', 0) if sps else 0)
            if available and qty > available:
                qty = available
            
            # IMMEDIATELY deduct from warehouse stock as "in transit" by increasing transfer_out
            if qty > 0:
                # Use an atomic update to reflect items sent but not yet validated
                SalesPointStock.objects.filter(pk=sps.pk).update(transfer_out=F('transfer_out') + qty)
                
                # Create stock transaction to track the deduction
                StockTransaction.objects.create(
                    salespoint=wh,
                    product_id=pid,
                    qty=-qty,  # Negative to show deduction
                    reason='restock_sent',
                    reference=ref,
                    user=request.user,
                )
                

            
            # Create restock line
            RestockLine.objects.create(
                request=restock_request,
                product_id=pid,
                quantity_requested=qty,
                quantity_approved=qty,  # Pre-approved by warehouse
            )
            created += 1
            
        # Create notification for the salespoint manager
        try:
            from django.contrib.auth import get_user_model
            User = get_user_model()
            from apps.sales.models import Notification
            
            # Find the salespoint manager
            manager = User.objects.filter(salespoint_id=to_sp, role__in=['sales_manager','gerant','gérant']).first()
            if manager:
                # Create detailed notification
                Notification.objects.create(
                    user=manager,
                    message=f"🚚 Nouvel approvisionnement reçu de l'entrepôt: {ref} ({created} produit(s))",
                    link=f"/sales/manager/inbound/",
                    kind="restock_incoming",
                )
        except Exception:
            # Silently handle notification errors
            pass
            
        # Notify warehouse managers that a restock has been sent
        try:
            from django.contrib.auth import get_user_model
            User = get_user_model()
            from apps.sales.models import Notification
            mgrs = User.objects.filter(role='warehouse_mgr', is_active=True)
            dest = None
            try:
                sp_obj = SalesPoint.objects.filter(id=to_sp).only('name').first()
                dest = sp_obj.name if sp_obj else 'Point de vente'
            except Exception:
                dest = 'Point de vente'
            for u in mgrs:
                Notification.objects.create(
                    user=u,
                    message=f"🚚 Approvisionnement expédié vers {dest}: {ref} ({created} produit(s))",
                    link="/inventory/warehouse/journal/",
                    kind="restock_sent",
                )
        except Exception:
            pass

    return JsonResponse({'ok': True, 'reference': ref, 'created': created, 'request_id': restock_request.id})


@login_required
def api_wh_salespoints(request):
    """List salespoints (excluding warehouse) with a hint of the manager name if available."""
    role = getattr(request.user, 'role', '')
    if not (request.user.is_superuser or role == 'warehouse_mgr' or getattr(request.user, 'is_staff', False)):
        return JsonResponse([], safe=False)
    rows = []
    for sp in SalesPoint.objects.filter(is_warehouse=False).order_by('name'):
        manager = ''
        try:
            from django.contrib.auth import get_user_model
            User = get_user_model()
            u = User.objects.filter(salespoint=sp, role__in=['sales_manager','gerant','gérant']).order_by('username').first()
            if u:
                manager = u.get_full_name() or u.username
        except Exception:
            pass
        rows.append({'id': sp.id, 'name': sp.name, 'manager': manager})
    return JsonResponse({'ok': True, 'rows': rows})


@login_required
def api_wh_ref(request):
    """Generate a WH reference like WH-DDMMYY-P-0001 or WH-...-M-...."""
    role = getattr(request.user, 'role', '')
    if not (request.user.is_superuser or role == 'warehouse_mgr' or getattr(request.user, 'is_staff', False)):
        return JsonResponse({'ok': False}, status=403)
    kind = (request.GET.get('kind') or 'P').upper()
    today = timezone.localdate()
    prefix = f"WH-{today.strftime('%d%m%y')}-{('M' if kind=='M' else 'P')}-"
    base = StockTransaction.objects.filter(reference__startswith=prefix).values('reference').annotate(c=Count('id'))
    used = 0
    for r in base:
        try:
            num = int(str(r['reference']).split('-')[-1])
        except Exception:
            num = 0
        used = max(used, num)
    seq = used + 1
    return JsonResponse({'ok': True, 'ref': f"{prefix}{seq:04d}"})


# ===== Warehouse Commande (purchase request to Commercial Director) =====

@login_required
def warehouse_commande(request):
    """Page alias for 'Commande des points de vente': show pending (sent) restock requests from salespoints.
    Redirects to the unified requests list filtered to status=sent.
    """
    role = getattr(request.user, 'role', '')
    if not (request.user.is_superuser or role == 'warehouse_mgr' or getattr(request.user, 'is_staff', False)):
        return redirect('sales:dashboard')
    # Use the existing requests list with default filters, enforce pending (sent)
    return redirect(f"{reverse('inventory:warehouse_requests')}?status=sent")


@login_required
def api_wh_cmd_ref(request):
    """Generate a CMD reference like CMD-WH-DDMMYY-0001"""
    role = getattr(request.user, 'role', '')
    if not (request.user.is_superuser or role == 'warehouse_mgr' or getattr(request.user, 'is_staff', False)):
        return JsonResponse({'ok': False}, status=403)
    today = timezone.localdate()
    prefix = f"CMD-WH-{today.strftime('%d%m%y')}-"
    # Count existing by scanning existing references starting with prefix
    used = 0
    for ref in WarehousePurchaseRequest.objects.filter(reference__startswith=prefix).values_list('reference', flat=True):
        try:
            num = int(str(ref).split('-')[-1])
        except Exception:
            num = 0
        used = max(used, num)
    seq = used + 1
    return JsonResponse({'ok': True, 'ref': f"{prefix}{seq:04d}"})


@login_required
def api_wh_cmd_submit(request):
    """Create a WarehousePurchaseRequest with lines.
    Body: { notes?: str, lines: [{product_id, qty}] }
    """
    role = getattr(request.user, 'role', '')
    if not (request.user.is_superuser or role == 'warehouse_mgr' or getattr(request.user, 'is_staff', False)):
        return JsonResponse({'ok': False, 'error': 'Accès refusé.'}, status=403)
    try:
        payload = json.loads(request.body.decode('utf-8'))
        notes = (payload.get('notes') or '').strip()
        lines = payload.get('lines') or []
    except Exception:
        return JsonResponse({'ok': False, 'error': 'Requête invalide.'}, status=400)
    if not isinstance(lines, list) or not lines:
        return JsonResponse({'ok': False, 'error': 'Aucun article.'}, status=400)

    today = timezone.localdate()
    prefix = f"CMD-WH-{today.strftime('%d%m%y')}-"
    with transaction.atomic():
        used = 0
        for ref in WarehousePurchaseRequest.objects.select_for_update().filter(reference__startswith=prefix).values_list('reference', flat=True):
            try:
                num = int(str(ref).split('-')[-1])
            except Exception:
                num = 0
            used = max(used, num)
        seq = used + 1
        ref = f"{prefix}{seq:04d}"

        req = WarehousePurchaseRequest.objects.create(
            requested_by=request.user,
            status='sent',
            reference=ref,
            notes=notes,
        )
        created = 0
        for ln in lines:
            pid = int(ln.get('product_id') or 0)
            qty = int(ln.get('qty') or 0)
            if pid and qty > 0:
                WarehousePurchaseLine.objects.create(
                    request=req,
                    product_id=pid,
                    quantity_requested=qty,
                )
                created += 1

    # Notify all warehouse managers about the sent CMD-WH
    try:
        from django.contrib.auth import get_user_model
        from apps.sales.models import Notification
        User = get_user_model()
        for u in User.objects.filter(role='warehouse_mgr', is_active=True):
            Notification.objects.create(
                user=u,
                message=f"📦 CMD-WH envoyée: {ref} ({created} ligne(s))",
                link="/admin/inventory/warehousepurchaserequest/",
                kind="cmd_wh_sent",
            )
    except Exception:
        pass

    return JsonResponse({'ok': True, 'id': req.id, 'reference': ref, 'created': created})


@login_required
def warehouse_request_print(request, req_id: int):
    if not (request.user.is_superuser or getattr(request.user, 'role', '') == 'warehouse_mgr' or getattr(request.user, 'is_staff', False)):
        return redirect('sales:dashboard')
    # Load header only (avoid heavy prefetch)
    req = get_object_or_404(
        RestockRequest.objects.select_related('salespoint', 'requested_by'),
        pk=req_id,
    )
    # Load lines efficiently
    lines_qs = (
        RestockLine.objects
        .select_related('product', 'product__brand')
        .filter(request_id=req.id)
        .only('product_id', 'quantity', 'quantity_requested', 'quantity_approved',
              'product__name', 'product__brand__name')
        .order_by('product__name')
    )
    return render(request, 'inventory/warehouse/warehouse_request_print.html', { 'req': req, 'lines': lines_qs })


# ===== Barcode Scanning APIs =====

@login_required
def api_scan_barcode(request):
    """Scan a barcode and return product/salespoint information."""
    # Only allow warehouse managers and superusers
    user_role = getattr(request.user, 'role', '')
    if not (request.user.is_superuser or user_role == 'warehouse_mgr'):
        return JsonResponse({'ok': False, 'error': 'Accès refusé.'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body.decode('utf-8'))
        barcode_data = data.get('barcode', '').strip()
        location_type = data.get('location_type', 'product')  # 'product' or 'salespoint'
    except Exception:
        return JsonResponse({'ok': False, 'error': 'Invalid request data'}, status=400)
    
    if not barcode_data:
        return JsonResponse({'ok': False, 'error': 'Barcode data required'}, status=400)
    
    try:
        if location_type == 'product':
            # Parse product barcode: PROD-{product_id}-{sku}-{name}
            if barcode_data.startswith('PROD-'):
                parts = barcode_data.split('-', 3)
                if len(parts) >= 2:
                    product_id = int(parts[1])
                    product = Product.objects.select_related('brand').get(id=product_id, is_active=True)
                    
                    # Get current stock at user's salespoint
                    user_sp = getattr(request.user, 'salespoint', None)
                    stock_info = None
                    if user_sp:
                        try:
                            stock = SalesPointStock.objects.get(salespoint=user_sp, product=product)
                            stock_info = {
                                'available_qty': stock.available_qty,
                                'opening_qty': stock.opening_qty,
                                'sold_qty': stock.sold_qty,
                                'transfer_in': stock.transfer_in,
                                'transfer_out': stock.transfer_out,
                                'reserved_qty': stock.reserved_qty,
                            }
                        except SalesPointStock.DoesNotExist:
                            stock_info = {'available_qty': 0}
                    
                    return JsonResponse({
                        'ok': True,
                        'type': 'product',
                        'product': {
                            'id': product.id,
                            'name': product.name,
                            'sku': product.sku,
                            'brand': getattr(product.brand, 'name', '') if product.brand else '',
                            'cost_price': str(product.cost_price or 0),
                            'retail_price': str(product.selling_price or 0),
                            'wholesale_price': str(product.wholesale_price or 0),
                        },
                        'stock': stock_info,
                    })
            
        elif location_type == 'salespoint':
            # Parse salespoint barcode: SP-{sp_id}-{name}
            if barcode_data.startswith('SP-'):
                parts = barcode_data.split('-', 2)
                if len(parts) >= 2:
                    sp_id = int(parts[1])
                    salespoint = SalesPoint.objects.get(id=sp_id)
                    
                    return JsonResponse({
                        'ok': True,
                        'type': 'salespoint',
                        'salespoint': {
                            'id': salespoint.id,
                            'name': salespoint.name,
                            'address': salespoint.address,
                            'phone': salespoint.phone,
                            'is_warehouse': salespoint.is_warehouse,
                        },
                    })
        
        return JsonResponse({'ok': False, 'error': 'Invalid barcode format'}, status=400)
        
    except Product.DoesNotExist:
        return JsonResponse({'ok': False, 'error': 'Product not found'}, status=404)
    except SalesPoint.DoesNotExist:
        return JsonResponse({'ok': False, 'error': 'Salespoint not found'}, status=404)
    except ValueError:
        return JsonResponse({'ok': False, 'error': 'Invalid barcode data'}, status=400)
    except Exception as e:
        return JsonResponse({'ok': False, 'error': f'Scan error: {str(e)}'}, status=500)


@login_required
def api_upload_proof_photo(request):
    """Upload a photo for proof of delivery/receipt."""
    # Only allow warehouse managers and superusers
    user_role = getattr(request.user, 'role', '')
    if not (request.user.is_superuser or user_role == 'warehouse_mgr'):
        return JsonResponse({'ok': False, 'error': 'Accès refusé.'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'Method not allowed'}, status=405)
    
    try:
        photo = request.FILES.get('photo')
        if not photo:
            return JsonResponse({'ok': False, 'error': 'No photo provided'}, status=400)
        
        # For now, we'll just return a placeholder URL
        # In production, you'd upload to cloud storage (AWS S3, etc.)
        photo_url = f"/media/proof_photos/{request.user.id}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.jpg"
        
        return JsonResponse({
            'ok': True,
            'photo_url': photo_url,
            'message': 'Photo uploaded successfully'
        })
        
    except Exception as e:
        return JsonResponse({'ok': False, 'error': f'Upload error: {str(e)}'}, status=500)


@login_required
def barcode_scanner(request):
    """Barcode scanner interface for products and salespoints."""
    # Only allow warehouse managers and superusers
    user_role = getattr(request.user, 'role', '')
    if not (request.user.is_superuser or user_role == 'warehouse_mgr'):
        return redirect('sales:dashboard')
    
    return render(request, 'inventory/warehouse/barcode_scanner.html', {})


@login_required
def barcode_printer(request):
    """Barcode printer interface for viewing and printing barcodes."""
    # Only allow warehouse managers and superusers
    user_role = getattr(request.user, 'role', '')
    if not (request.user.is_superuser or user_role == 'warehouse_mgr'):
        return redirect('sales:dashboard')
    
    return render(request, 'inventory/warehouse/barcode_printer.html', {})


@login_required
def api_barcode_list(request):
    """API to list available barcodes for printing."""
    user_role = getattr(request.user, 'role', '')
    if not (request.user.is_superuser or user_role == 'warehouse_mgr'):
        return JsonResponse({'ok': False, 'error': 'Accès refusé.'}, status=403)
    
    import os
    from pathlib import Path
    
    barcodes_dir = Path('static/barcodes')
    barcodes = []
    
    if barcodes_dir.exists():
        for file_path in barcodes_dir.glob('*.png'):
            filename = file_path.name
            # Extract product info from filename
            if filename.startswith('prod_'):
                parts = filename.replace('.png', '').split('_', 2)
                if len(parts) >= 3:
                    product_id = parts[1]
                    product_name = parts[2].replace('_', ' ')
                    barcodes.append({
                        'filename': filename,
                        'name': product_name,
                        'type': 'product',
                        'id': product_id
                    })
            elif filename.startswith('sp_'):
                parts = filename.replace('.png', '').split('_', 2)
                if len(parts) >= 3:
                    sp_id = parts[1]
                    sp_name = parts[2].replace('_', ' ')
                    barcodes.append({
                        'filename': filename,
                        'name': sp_name,
                        'type': 'salespoint',
                        'id': sp_id
                    })
    
    return JsonResponse({
        'ok': True,
        'barcodes': sorted(barcodes, key=lambda x: x['name']),
        'total': len(barcodes)
    })


@login_required
def api_generate_barcodes(request):
    """API to generate new barcodes."""
    user_role = getattr(request.user, 'role', '')
    if not (request.user.is_superuser or user_role == 'warehouse_mgr'):
        return JsonResponse({'ok': False, 'error': 'Accès refusé.'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'Method not allowed'}, status=405)
    
    try:
        # Run the barcode generation command
        import subprocess
        result = subprocess.run([
            'python', 'manage.py', 'generate_barcodes',
            '--output-dir', 'static/barcodes',
            '--format', 'qr',
            '--size', '150'
        ], capture_output=True, text=True, cwd='.')
        
        if result.returncode == 0:
            return JsonResponse({
                'ok': True,
                'message': 'Codes-barres générés avec succès',
                'output': result.stdout
            })
        else:
            return JsonResponse({
                'ok': False,
                'error': 'Erreur lors de la génération',
                'details': result.stderr
            })
    except Exception as e:
        return JsonResponse({
            'ok': False,
            'error': f'Erreur: {str(e)}'
        })


@login_required
def api_wh_restock_validate(request, req_id: int):
    """Validate a restock request by moving stock from the warehouse to the salespoint.
    This deducts from warehouse stock and adds to the salespoint stock.
    """
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'Method not allowed'}, status=405)
    if not (request.user.is_superuser or getattr(request.user, 'role', '') == 'warehouse_mgr' or getattr(request.user, 'is_staff', False)):
        return JsonResponse({'ok': False, 'error': 'Accès refusé.'}, status=403)

    req = get_object_or_404(
        RestockRequest.objects.select_related('salespoint').prefetch_related('lines__product'),
        pk=req_id,
    )

    # Identify warehouse salespoint
    warehouse_sp = SalesPoint.objects.filter(is_warehouse=True).first()
    if not warehouse_sp:
        return JsonResponse({'ok': False, 'error': "Le point 'Entrepôt' n'est pas configuré."}, status=400)

    # Optional: limit to selected lines sent by client
    selected_ids = None
    try:
        payload = json.loads(request.body.decode('utf-8')) if request.body else {}
        if isinstance(payload, dict) and isinstance(payload.get('lines'), list):
            # Expect [{product_id, qty?}] but qty is ignored here; we use request quantities
            selected_ids = {int(x.get('product_id')) for x in payload['lines'] if str(x.get('product_id') or '').isdigit()}
            if not selected_ids:
                selected_ids = None
    except Exception:
        selected_ids = None

    # Simple availability check and movement
    moved = []
    with transaction.atomic():
        for ln in req.lines.select_related('product').all():
            if selected_ids is not None and ln.product_id not in selected_ids:
                continue
            qty = int(ln.quantity_approved or ln.quantity_requested or ln.quantity or 0)
            if qty <= 0:
                continue
            # Update destination salespoint stock (transfer_in += qty)
            sps_dest, _ = SalesPointStock.objects.select_for_update().get_or_create(
                salespoint=req.salespoint,
                product=ln.product,
                defaults={'opening_qty': 0},
            )
            SalesPointStock.objects.filter(pk=sps_dest.pk).update(transfer_in=F('transfer_in') + qty)

            # Update warehouse stock (transfer_out += qty)
            sps_wh, _ = SalesPointStock.objects.select_for_update().get_or_create(
                salespoint=warehouse_sp,
                product=ln.product,
                defaults={'opening_qty': 0},
            )
            SalesPointStock.objects.filter(pk=sps_wh.pk).update(transfer_out=F('transfer_out') + qty)

            # Optional audit transactions
            StockTransaction.create_transaction(
                salespoint=warehouse_sp,
                product=ln.product,
                qty=-qty,
                reason='restock',
                reference=req.reference or f"REQ{req.id}",
                user=request.user,
                document_type='RestockRequest',
                document_id=req.id,
            )
            StockTransaction.create_transaction(
                salespoint=req.salespoint,
                product=ln.product,
                qty=qty,
                reason='restock',
                reference=req.reference or f"REQ{req.id}",
                user=request.user,
                document_type='RestockRequest',
                document_id=req.id,
            )

            moved.append({'product_id': ln.product_id, 'qty': qty})

        # Update request status
        req.status = 'validated'
        req.validated_at = timezone.now()
        req.save(update_fields=['status', 'validated_at', 'updated_at'])

    return JsonResponse({'ok': True, 'moved': moved, 'request_id': req.id, 'status': req.status})


# ===== Historiques des transferts (entre points de vente) =====
@login_required
def transfer_history(request):
    if not (request.user.is_superuser or getattr(request.user, 'role', '') == 'warehouse_mgr' or getattr(request.user, 'is_staff', False)):
        return redirect('sales:dashboard')

    q = (request.GET.get('q') or '').strip()
    df = (request.GET.get('from') or '').strip()
    dt = (request.GET.get('to') or '').strip()
    sp_from = int(request.GET.get('from_sp') or 0)
    sp_to = int(request.GET.get('to_sp') or 0)
    status = (request.GET.get('status') or '').strip()

    qs = TransferRequest.objects.select_related('from_salespoint','to_salespoint','requested_by').order_by('-created_at')
    if q:
        qs = qs.filter(
            Q(from_salespoint__name__icontains=q) |
            Q(to_salespoint__name__icontains=q) |
            Q(requested_by__username__icontains=q) |
            Q(lines__product__name__icontains=q) |
            Q(lines__product__brand__name__icontains=q)
        ).distinct()
    if df:
        qs = qs.filter(created_at__date__gte=df)
    if dt:
        qs = qs.filter(created_at__date__lte=dt)
    if sp_from:
        qs = qs.filter(from_salespoint_id=sp_from)
    if sp_to:
        qs = qs.filter(to_salespoint_id=sp_to)
    if status:
        qs = qs.filter(status=status)

    # Totals across filtered set
    totals_qs = TransferRequestLine.objects.filter(request__in=qs)
    # TransferRequestLine has no quantity_approved field; use quantity for both totals
    totals = totals_qs.aggregate(total_requested=Sum('quantity'))
    totals['total_approved'] = totals.get('total_requested') or 0

    # Pagination
    try:
        page_num = int(request.GET.get('page') or 1)
    except Exception:
        page_num = 1
    try:
        per_page = int(request.GET.get('per_page') or 50)
        if per_page not in [25, 50, 100, 200]:
            per_page = 50
    except Exception:
        per_page = 50
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    paginator = Paginator(qs, per_page)
    try:
        page = paginator.page(page_num)
    except PageNotAnInteger:
        page = paginator.page(1)
    except EmptyPage:
        page = paginator.page(paginator.num_pages)

    # Salespoints list for filters
    salespoints = SalesPoint.objects.filter(is_warehouse=False).order_by('name')

    return render(request, 'inventory/warehouse/transfer_history.html', {
        'rows': page.object_list,
        'paginator': paginator,
        'page': page,
        'q': q,
        'date_from': df,
        'date_to': dt,
        'sp_from': sp_from,
        'sp_to': sp_to,
        'status': status,
        'salespoints': salespoints,
        'total_requested': totals.get('total_requested') or 0,
        'total_approved': totals.get('total_approved') or 0,
    })


@login_required
def api_transfer_request_lines(request, req_id: int):
    if not (request.user.is_superuser or getattr(request.user, 'role', '') == 'warehouse_mgr' or getattr(request.user, 'is_staff', False)):
        return JsonResponse({"ok": False, "error": "Accès refusé."}, status=403)

    req = get_object_or_404(
        TransferRequest.objects.select_related('from_salespoint','to_salespoint','requested_by'),
        pk=req_id,
    )
    lines_qs = (
        TransferRequestLine.objects
        .select_related('product','product__brand')
        .filter(request_id=req.id)
        .only('product_id','quantity','product__name','product__brand__name')
        .order_by('product__name')
    )
    approved_by = getattr(req, 'approved_by', None)
    approved_by_name = getattr(approved_by, 'username', None) if approved_by else None
    approved_at = getattr(req, 'approved_at', None)
    data = {
        'ok': True,
        'id': req.id,
        'from': getattr(req.from_salespoint, 'name', ''),
        'to': getattr(req.to_salespoint, 'name', ''),
        'requested_by': getattr(req.requested_by, 'username', ''),
        'approved_by': approved_by_name,
        'approved_at': approved_at.strftime('%d/%m/%Y %H:%M') if approved_at else None,
        'created_at': req.created_at.strftime('%d/%m/%Y %H:%M') if req.created_at else '',
        'status': getattr(req, 'status', ''),
        'lines': [
            {
                'product_id': ln.product_id,
                'name': getattr(ln.product, 'name', f"#{ln.product_id}"),
                'brand': getattr(getattr(ln.product, 'brand', None), 'name', ''),
                'qty_requested': int(getattr(ln, 'quantity', 0) or 0),
                'qty_approved': int(getattr(ln, 'quantity_approved', getattr(ln, 'quantity', 0)) or 0),
            }
            for ln in lines_qs
        ],
    }
    return JsonResponse(data)


@login_required
def transfer_history_export_csv(request):
    if not (request.user.is_superuser or getattr(request.user, 'role', '') == 'warehouse_mgr' or getattr(request.user, 'is_staff', False)):
        return redirect('sales:dashboard')
    q = (request.GET.get('q') or '').strip()
    df = (request.GET.get('from') or '').strip()
    dt = (request.GET.get('to') or '').strip()
    sp_from = int(request.GET.get('from_sp') or 0)
    sp_to = int(request.GET.get('to_sp') or 0)
    status = (request.GET.get('status') or '').strip()

    qs = TransferRequest.objects.select_related('from_salespoint','to_salespoint','requested_by').order_by('-created_at')
    if q:
        qs = qs.filter(
            Q(from_salespoint__name__icontains=q) |
            Q(to_salespoint__name__icontains=q) |
            Q(requested_by__username__icontains=q) |
            Q(lines__product__name__icontains=q) |
            Q(lines__product__brand__name__icontains=q)
        ).distinct()
    if df:
        qs = qs.filter(created_at__date__gte=df)
    if dt:
        qs = qs.filter(created_at__date__lte=dt)
    if sp_from:
        qs = qs.filter(from_salespoint_id=sp_from)
    if sp_to:
        qs = qs.filter(to_salespoint_id=sp_to)
    if status:
        qs = qs.filter(status=status)

    # Build CSV
    import csv
    from django.http import HttpResponse
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="transfer_history.csv"'
    writer = csv.writer(response)
    writer.writerow(['Date','De','Vers','Demandeur','Statut','Produit','Marque','Qté demandée','Qté approuvée'])
    lines = TransferRequestLine.objects.select_related('request','product','product__brand','request__from_salespoint','request__to_salespoint','request__requested_by').filter(request__in=qs).order_by('-request__created_at','product__name')
    for ln in lines:
        req = ln.request
        writer.writerow([
            req.created_at.strftime('%d/%m/%Y %H:%M') if req.created_at else '',
            getattr(req.from_salespoint,'name',''),
            getattr(req.to_salespoint,'name',''),
            getattr(req.requested_by,'username',''),
            getattr(req,'status',''),
            getattr(ln.product,'name',f'#{ln.product_id}'),
            getattr(getattr(ln.product,'brand',None),'name',''),
            int(getattr(ln,'quantity',0) or 0),
            int(getattr(ln,'quantity',0) or 0),
        ])
    return response


@login_required
def restock_stats(request):
    if not (request.user.is_superuser or getattr(request.user, 'role', '') == 'warehouse_mgr' or getattr(request.user, 'is_staff', False)):
        return redirect('sales:dashboard')

    sp_id = int(request.GET.get('sp') or 0)
    prod_q = (request.GET.get('product') or '').strip()
    df = (request.GET.get('from') or '').strip()
    dt = (request.GET.get('to') or '').strip()
    
    # Default to today if no date filters provided
    today = timezone.localdate()
    if not df and not dt:
        df = today.strftime('%Y-%m-%d')
        dt = today.strftime('%Y-%m-%d')

    qs = RestockRequest.objects.exclude(reference__startswith='WH-RQ-').filter(status__in=['validated','partially_validated'])
    if sp_id:
        qs = qs.filter(salespoint_id=sp_id)
    if df:
        qs = qs.filter(created_at__date__gte=df)
    if dt:
        qs = qs.filter(created_at__date__lte=dt)

    # Lines joined to products; use approved quantity if present else requested
    lines = RestockLine.objects.select_related('product','product__brand','request').filter(request__in=qs)
    if prod_q:
        lines = lines.filter(Q(product__name__icontains=prod_q) | Q(product__brand__name__icontains=prod_q))

    # Build rows grouped by product then date
    from collections import defaultdict
    rows = []
    grouped = defaultdict(lambda: defaultdict(int))
    meta = {}
    for ln in lines.only('product_id','quantity_requested','quantity_approved','request__created_at','product__name','product__brand__name'):
        date_key = ln.request.created_at.date() if ln.request and ln.request.created_at else None
        if not date_key:
            continue
        qty = int(ln.quantity_approved or ln.quantity_requested or 0)
        grouped[ln.product_id][date_key] += qty
        if ln.product_id not in meta:
            meta[ln.product_id] = {
                'name': getattr(ln.product, 'name', f'#{ln.product_id}'),
                'brand': getattr(getattr(ln.product,'brand',None),'name',''),
            }

    # Flatten for template
    for pid, date_map in grouped.items():
        product_info = meta.get(pid, {'name': f'#{pid}', 'brand': ''})
        for d, qsum in sorted(date_map.items()):
            rows.append({
                'product_id': pid,
                'name': product_info['name'],
                'brand': product_info['brand'],
                'date': d,
                'qty': qsum,
            })

    # Totals
    total_qty = sum(r['qty'] for r in rows)

    # Salespoints for filter
    salespoints = SalesPoint.objects.filter(is_warehouse=False).order_by('name')

    # Pagination
    try:
        page_num = int(request.GET.get('page') or 1)
    except Exception:
        page_num = 1
    try:
        per_page = int(request.GET.get('per_page') or 100)
        if per_page not in [50, 100, 200, 500]:
            per_page = 100
    except Exception:
        per_page = 100
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    paginator = Paginator(rows, per_page)
    try:
        page = paginator.page(page_num)
    except PageNotAnInteger:
        page = paginator.page(1)
    except EmptyPage:
        page = paginator.page(paginator.num_pages)

    return render(request, 'inventory/warehouse/restock_stats.html', {
        'rows': page.object_list,
        'paginator': paginator,
        'page': page,
        'salespoints': salespoints,
        'sp_id': sp_id,
        'product_q': prod_q,
        'date_from': df,
        'date_to': dt,
        'total_qty': total_qty,
    })


@login_required
def restock_stats_export_csv(request):
    if not (request.user.is_superuser or getattr(request.user, 'role', '') == 'warehouse_mgr' or getattr(request.user, 'is_staff', False)):
        return redirect('sales:dashboard')
    sp_id = int(request.GET.get('sp') or 0)
    prod_q = (request.GET.get('product') or '').strip()
    df = (request.GET.get('from') or '').strip()
    dt = (request.GET.get('to') or '').strip()
    
    # Default to today if no date filters provided (same logic as main view)
    today = timezone.localdate()
    if not df and not dt:
        df = today.strftime('%Y-%m-%d')
        dt = today.strftime('%Y-%m-%d')

    qs = RestockRequest.objects.exclude(reference__startswith='WH-RQ-').filter(status__in=['validated','partially_validated'])
    if sp_id:
        qs = qs.filter(salespoint_id=sp_id)
    if df:
        qs = qs.filter(created_at__date__gte=df)
    if dt:
        qs = qs.filter(created_at__date__lte=dt)
    lines = RestockLine.objects.select_related('product','product__brand','request').filter(request__in=qs)
    if prod_q:
        lines = lines.filter(Q(product__name__icontains=prod_q) | Q(product__brand__name__icontains=prod_q))

    import csv
    from django.http import HttpResponse
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="restock_stats.csv"'
    writer = csv.writer(response)
    writer.writerow(['Salespoint','Date','Produit','Marque','Qté'])
    for ln in lines.order_by('request__salespoint__name','product__name','request__created_at'):
        d = ln.request.created_at.strftime('%d/%m/%Y') if ln.request and ln.request.created_at else ''
        writer.writerow([
            getattr(getattr(ln.request,'salespoint',None),'name',''),
            d,
            getattr(ln.product,'name',f'#{ln.product_id}'),
            getattr(getattr(ln.product,'brand',None),'name',''),
            int(ln.quantity_approved or ln.quantity_requested or 0),
        ])
    return response